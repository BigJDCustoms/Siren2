
from flask import Flask, render_template, request, redirect, send_file
import os
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def get_top_level_group(base_folder, file_path):
    rel_path = os.path.relpath(file_path, base_folder)
    parts = rel_path.split(os.sep)
    return parts[0] if parts else ""

def extract_siren_data_from_all_folders(base_folder):
    entries = []
    for root, dirs, files in os.walk(base_folder):
        for file in files:
            if file.lower() == "carvariations.meta":
                file_path = os.path.join(root, file)
                try:
                    tree = ET.parse(file_path)
                    root_element = tree.getroot()
                    for item in root_element.iter():
                        model = None
                        siren = None
                        for elem in item.iter():
                            if elem.tag == "modelName":
                                model = (elem.text or "").strip()
                            if elem.tag == "sirenSettings" and "value" in elem.attrib:
                                siren = elem.attrib["value"]
                        if model and siren and siren.strip() != "0":
                            group = get_top_level_group(base_folder, file_path)
                            entries.append((model, siren, group))
                except ET.ParseError:
                    continue
    return entries

def find_conflicts(data):
    siren_map = defaultdict(list)
    for model, siren, group in data:
        siren_map[siren].append((model, group))
    conflicts = {
        sid: models for sid, models in siren_map.items()
        if len(set(group for _, group in models)) > 1
    }
    return conflicts

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        uploaded_file = request.files["zip_file"]
        if uploaded_file.filename.endswith(".zip"):
            filename = secure_filename(uploaded_file.filename)
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            uploaded_file.save(file_path)

            extract_path = os.path.join(app.config["UPLOAD_FOLDER"], "extracted")
            if os.path.exists(extract_path):
                shutil.rmtree(extract_path)
            os.makedirs(extract_path)

            with zipfile.ZipFile(file_path, "r") as zip_ref:
                zip_ref.extractall(extract_path)

            results = extract_siren_data_from_all_folders(extract_path)
            conflicts = find_conflicts(results)

            wb = Workbook()
            ws = wb.active
            ws.title = "Siren Settings"
            ws.append(["Model Name", "Siren ID", "Pack/Group"])

            red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

            for model, siren, group in results:
                is_conflict = siren in conflicts
                row = [model, siren, group]
                ws.append(row)
                if is_conflict:
                    for cell in ws[ws.max_row]:
                        cell.fill = red_fill

            output_file = os.path.join(app.config["UPLOAD_FOLDER"], "siren_conflicts_results.xlsx")
            wb.save(output_file)

            return send_file(output_file, as_attachment=True)

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
