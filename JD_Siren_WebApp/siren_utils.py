import os
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

def extract_zip(zip_path, extract_to):
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)

def get_top_level_group(base_folder, file_path):
    rel_path = os.path.relpath(file_path, base_folder)
    parts = rel_path.split(os.sep)
    return parts[0] if parts else ""

def extract_siren_data_from_all_folders(base_folder):
    entries = []
    for root, dirs, files in os.walk(base_folder):
        for file in files:
            if file.lower() == "carvariations.meta":
                file_path = os.path.join(root, file)
                try:
                    tree = ET.parse(file_path)
                    root_element = tree.getroot()
                    for item in root_element.iter():
                        model = None
                        siren = None
                        for elem in item.iter():
                            if elem.tag == "modelName":
                                model = (elem.text or "").strip()
                            if elem.tag == "sirenSettings" and "value" in elem.attrib:
                                siren = elem.attrib["value"].strip()
                        if model and siren and siren != "0":
                            group = get_top_level_group(base_folder, file_path)
                            entries.append((model, siren, group))
                except ET.ParseError:
                    continue
    return entries

def find_conflicts(data):
    siren_map = defaultdict(list)
    for model, siren, group in data:
        siren_map[siren].append((model, group))
    conflicts = {
        sid: models for sid, models in siren_map.items()
        if len(set(group for _, group in models)) > 1
    }
    return conflicts

def process_zip(zip_path, result_dir):
    extract_dir = os.path.join(result_dir, 'unzipped')
    if os.path.exists(extract_dir):
        shutil.rmtree(extract_dir)
    os.makedirs(extract_dir, exist_ok=True)
    extract_zip(zip_path, extract_dir)

    results = extract_siren_data_from_all_folders(extract_dir)
    conflicts = find_conflicts(results)

    # TXT export
    txt_path = os.path.join(result_dir, "siren_conflicts_results.txt")
    with open(txt_path, "w") as f:
        for model, siren, group in results:
            f.write(f"{model}: {siren} (in {group})\n")
        if conflicts:
            f.write("\n--- Conflicts Detected ---\n")
            for sid, models in conflicts.items():
                f.write(f"Siren ID {sid} used by:\n")
                for m, g in models:
                    f.write(f"  - {m} (in {g})\n")
                f.write("\n")

    # XLSX export
    xlsx_path = os.path.join(result_dir, "siren_conflicts_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Siren Settings"
    ws.append(["Model Name", "Siren ID", "Pack/Group"])

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for model, siren, group in results:
        is_conflict = siren in conflicts
        row = [model, siren, group]
        ws.append(row)
        if is_conflict:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill

    wb.save(xlsx_path)

    return "siren_conflicts_results.txt", "siren_conflicts_results.xlsx"
